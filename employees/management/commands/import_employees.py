import time
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from employees.models import Employee
from employees.utils import geocode_address


class Command(BaseCommand):
    help = 'Import employees from Excel and geocode addresses via Yandex API'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            default='Docs/Kitap1.xlsx',
            help='Path to the Excel file (default: Docs/Kitap1.xlsx)',
        )
        parser.add_argument(
            '--delay',
            type=float,
            default=0.15,
            help='Seconds between Yandex API calls (default: 0.15)',
        )
        parser.add_argument(
            '--output',
            default='Docs/geocode_results.xlsx',
            help='Output Excel file path (default: Docs/geocode_results.xlsx)',
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options['file'])
        if not xlsx_path.exists():
            self.stderr.write(self.style.ERROR(f'File not found: {xlsx_path}'))
            return

        delay = options['delay']
        output_path = Path(options['output'])

        wb_in = openpyxl.load_workbook(xlsx_path)
        ws_in = wb_in.active
        rows = list(ws_in.iter_rows(min_row=2, values_only=True))
        total = len(rows)
        self.stdout.write(f'Found {total} rows. Starting import...')

        results = []
        ok_count = 0
        failed_count = 0
        skipped_count = 0

        for i, row in enumerate(rows, start=1):
            if not row or not row[0]:
                continue
            personnel_code = str(row[0]).strip()
            address = str(row[1]).strip() if row[1] else ''

            emp, created = Employee.objects.get_or_create(
                personnel_code=personnel_code,
                defaults={'address': address, 'geocode_status': 'pending'},
            )

            if not created and emp.geocode_status == 'ok':
                self.stdout.write(f'[{i}/{total}] SKIP {personnel_code} (already geocoded)')
                skipped_count += 1
                results.append({
                    'P': personnel_code,
                    'ADRES': emp.address,
                    'API_ADRES': emp.api_address,
                    'LAT': emp.lat,
                    'LNG': emp.lng,
                    'STATUS': 'ok',
                })
                continue

            geo = geocode_address(address)
            if geo['ok']:
                emp.lat = geo['lat']
                emp.lng = geo['lng']
                emp.api_address = geo['api_address']
                emp.geocode_status = 'ok'
                emp.address = address
                emp.save()
                ok_count += 1
                self.stdout.write(f'[{i}/{total}] OK  {personnel_code}')
                results.append({
                    'P': personnel_code,
                    'ADRES': address,
                    'API_ADRES': geo['api_address'],
                    'LAT': geo['lat'],
                    'LNG': geo['lng'],
                    'STATUS': 'ok',
                })
            else:
                emp.geocode_status = 'failed'
                emp.address = address
                emp.save()
                failed_count += 1
                self.stdout.write(self.style.WARNING(f'[{i}/{total}] FAIL {personnel_code}'))
                results.append({
                    'P': personnel_code,
                    'ADRES': address,
                    'API_ADRES': '',
                    'LAT': '',
                    'LNG': '',
                    'STATUS': 'failed',
                })

            time.sleep(delay)

        # --- Write output xlsx ---
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = 'Geocode Results'
        ws_out.append(['P', 'ADRES', 'API_ADRES', 'LAT', 'LNG', 'STATUS'])
        for r in results:
            ws_out.append([r['P'], r['ADRES'], r['API_ADRES'], r['LAT'], r['LNG'], r['STATUS']])

        # Auto-width columns
        for col in ws_out.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        wb_out.save(output_path)

        self.stdout.write(self.style.SUCCESS(
            f'\nDone! OK: {ok_count}  Failed: {failed_count}  Skipped: {skipped_count}'
        ))
        self.stdout.write(self.style.SUCCESS(f'Results saved to: {output_path}'))
