import math
import time
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from employees.utils import geocode_address

DOCS = settings.BASE_DIR / 'Docs'
INPUT_FILE = DOCS / 'Birey_Listesi.xlsx'
YANDEX_FILE = DOCS / 'geocode_results.xlsx'
OUTPUT_FILE = DOCS / 'geocode_comparison.xlsx'

DIFF_THRESHOLD_M = 500  # Bu metrenin üzerindeki farklar "farklı" sayılır


def haversine_m(lat1, lng1, lat2, lng2):
    R = 6_371_000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2
         + math.cos(lat1 * p) * math.cos(lat2 * p)
         * math.sin((lng2 - lng1) * p / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))


class Command(BaseCommand):
    help = 'Birey_Listesi adreslerini Mapbox ile geocode eder, Yandex sonuçlarıyla karşılaştırır'

    def handle(self, *args, **kwargs):
        # --- Yandex sonuçlarını yükle ---
        yandex = {}
        wb_y = openpyxl.load_workbook(YANDEX_FILE)
        for row in wb_y.active.iter_rows(min_row=2, values_only=True):
            p, adres, api_adres, lat, lng, status = row
            yandex[p] = {'lat': lat, 'lng': lng, 'api_adres': api_adres, 'status': status}

        # --- Birey listesini yükle ---
        wb_in = openpyxl.load_workbook(INPUT_FILE)
        employees = list(wb_in.active.iter_rows(min_row=2, values_only=True))

        total = len(employees)
        self.stdout.write(f'{total} çalışan geocode edilecek...\n')

        # --- Çıktı dosyasını hazırla ---
        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = 'Karşılaştırma'
        ws.append([
            'P', 'ADRES',
            'YANDEX_LAT', 'YANDEX_LNG', 'YANDEX_API_ADRES',
            'MAPBOX_LAT', 'MAPBOX_LNG', 'MAPBOX_API_ADRES',
            'FARK_M', 'MAPBOX_STATUS', 'FARKLI_MI',
        ])

        diff_count = 0
        fail_count = 0

        for i, (p_code, address) in enumerate(employees, 1):
            result = geocode_address(address)
            time.sleep(0.05)  # Rate limit

            yandex_row = yandex.get(p_code, {})
            y_lat = yandex_row.get('lat')
            y_lng = yandex_row.get('lng')
            y_api = yandex_row.get('api_adres', '')

            if result['ok']:
                m_lat = result['lat']
                m_lng = result['lng']
                m_api = result['api_address']
                m_status = 'ok'

                if y_lat and y_lng:
                    dist = haversine_m(y_lat, y_lng, m_lat, m_lng)
                    farkli = 'EVET' if dist > DIFF_THRESHOLD_M else 'hayır'
                    if dist > DIFF_THRESHOLD_M:
                        diff_count += 1
                else:
                    dist = None
                    farkli = 'yandex_yok'
            else:
                m_lat = m_lng = m_api = None
                m_status = result.get('reason', 'failed')
                dist = None
                farkli = 'mapbox_failed'
                fail_count += 1

            ws.append([
                p_code, address,
                y_lat, y_lng, y_api,
                m_lat, m_lng, m_api,
                round(dist) if dist is not None else None,
                m_status, farkli,
            ])

            if i % 50 == 0:
                self.stdout.write(f'  {i}/{total}...')

        # Farklı satırları sarıya boya
        yellow = openpyxl.styles.PatternFill('solid', fgColor='FFFF00')
        red = openpyxl.styles.PatternFill('solid', fgColor='FF9999')
        for row in ws.iter_rows(min_row=2):
            farkli_val = row[10].value
            if farkli_val == 'EVET':
                for cell in row:
                    cell.fill = yellow
            elif farkli_val == 'mapbox_failed':
                for cell in row:
                    cell.fill = red

        wb_out.save(OUTPUT_FILE)
        self.stdout.write(self.style.SUCCESS(
            f'\nTamamlandı → {OUTPUT_FILE}\n'
            f'  Farklı (>{DIFF_THRESHOLD_M}m): {diff_count}\n'
            f'  Mapbox başarısız: {fail_count}\n'
            f'  Toplam: {total}'
        ))
