import io
import logging
import threading
import time

import openpyxl

logger = logging.getLogger(__name__)
from django.http import HttpResponse
from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from workspaces.models import Workspace
from .models import Employee, ImportJob
from .serializers import EmployeeSerializer, EmployeeCreateSerializer
from .utils import geocode_address


def _run_import_job(job_id, rows, force):
    from django.db import connection
    connection.close()  # thread'e yeni DB bağlantısı açtır

    try:
        job = ImportJob.objects.get(id=job_id)
    except ImportJob.DoesNotExist:
        return

    for row in rows:
        job.refresh_from_db(fields=['stop_requested'])
        if job.stop_requested:
            job.log.append({'color': '#f0a030', 'text': f'⏹ Durduruldu ({job.done}/{job.total})'})
            job.status = 'stopped'
            job.save(update_fields=['log', 'status', 'updated_at'])
            return

        personnel_code = str(row.get('code', '')).strip()
        address = str(row.get('address', '')).strip().lower()
        job.log.append({'color': '#555', 'text': f'[{job.done}/{job.total}] {personnel_code} → {address}'})

        if not force and row.get('exists') and row.get('existing_status') == 'ok':
            job.done += 1
            job.skipped += 1
            job.log.append({'color': '#555', 'text': '  – atlandı (API çağrısı yapılmadı)'})
            job.save(update_fields=['done', 'skipped', 'log', 'updated_at'])
            continue

        logger.info('import_job [%s]: geocode → "%s"', personnel_code, address)
        result = geocode_address(address)
        emp, _ = Employee.objects.get_or_create(
            personnel_code=personnel_code,
            defaults={'address': address, 'geocode_status': 'pending'},
        )

        if result['ok']:
            emp.lat, emp.lng = result['lat'], result['lng']
            emp.api_address = result['api_address']
            emp.geocode_status = 'ok'
            emp.address = address
            emp.save()
            job.ok += 1
            job.log.append({'color': '#4cde8a', 'text': f'  ✓ ok → {result["api_address"]}'})
            logger.info('import_job [%s]: OK → %s', personnel_code, result['api_address'])
        else:
            emp.geocode_status = 'failed'
            emp.address = address
            emp.save()
            job.failed += 1
            reason = result.get('reason', 'unknown')
            job.errors.append({'code': personnel_code, 'reason': reason})
            job.log.append({'color': '#e05a5a', 'text': f'  ✗ hata: {reason}'})
            logger.error('import_job [%s]: HATA %s', personnel_code, reason)

        job.done += 1
        job.save(update_fields=['done', 'ok', 'failed', 'log', 'errors', 'updated_at'])
        time.sleep(0.18)

    job.status = 'done'
    job.save(update_fields=['status', 'updated_at'])


def manage_view(request):
    return render(request, 'employees/manage.html')


def sample_excel(_request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Personel'
    ws.append(['Sicil No', 'Adres'])
    ws.append(['W001', 'Soğanlı, İstanbul Cd No:343-345, 16150 Osmangazi/Bursa'])
    ws.append(['W002', 'Ataevler, Ali Rıza Bey Cd. No:47, 16210 Nilüfer/Bursa'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="ornek_personel.xlsx"'
    return response


class EmployeeViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeeSerializer
    http_method_names = ['get', 'post', 'patch', 'delete', 'head', 'options']

    def get_queryset(self):
        if self.action != 'list':
            return Employee.objects.all()
        show_all = self.request.query_params.get('all') == '1'
        qs = Employee.objects.all() if show_all else Employee.objects.filter(geocode_status='ok')
        workspace_id = self.request.query_params.get('workspace')
        if workspace_id:
            qs = qs.filter(workspaces__id=workspace_id)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = EmployeeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch'], url_path='update-location')
    def update_location(self, request, pk=None):
        try:
            employee = Employee.objects.get(pk=pk)
        except Employee.DoesNotExist:
            return Response({'error': 'Çalışan bulunamadı.'}, status=404)
        lat = request.data.get('lat')
        lng = request.data.get('lng')
        address = request.data.get('address')
        api_address = request.data.get('api_address')
        if lat is None or lng is None:
            return Response({'error': 'lat ve lng zorunludur.'}, status=400)
        employee.lat = float(lat)
        employee.lng = float(lng)
        if address is not None:
            employee.address = address
        if api_address is not None:
            employee.api_address = api_address
        employee.geocode_status = 'ok'
        employee.save()
        return Response(EmployeeSerializer(employee).data)

    @action(detail=False, methods=['post'], url_path='geocode')
    def geocode(self, request):
        address = request.data.get('address', '').strip()
        if not address:
            return Response({'error': 'Adres alanı boş olamaz.'}, status=400)
        result = geocode_address(address)
        if not result['ok']:
            if result['reason'] == 'no_result':
                return Response(
                    {'error': 'Adres bulunamadı. Daha ayrıntılı bir adres deneyin.'},
                    status=422,
                )
            return Response(
                {'error': 'Yandex API hatası: ' + result.get('detail', '')},
                status=502,
            )
        return Response({'lat': result['lat'], 'lng': result['lng'], 'api_address': result['api_address']})

    @action(detail=False, methods=['post'], url_path='import-start')
    def import_start(self, request):
        rows = request.data.get('rows', [])
        force = request.data.get('force', False)
        if not rows:
            return Response({'error': 'Satır listesi boş'}, status=400)
        job = ImportJob.objects.create(total=len(rows))
        threading.Thread(target=_run_import_job, args=(job.id, rows, force), daemon=True).start()
        return Response({'job_id': str(job.id)})

    @action(detail=False, methods=['get'], url_path='import-status')
    def import_status(self, request):
        job_id = request.query_params.get('job_id')
        if not job_id:
            return Response({'error': 'job_id gerekli'}, status=400)
        try:
            job = ImportJob.objects.get(id=job_id)
        except ImportJob.DoesNotExist:
            return Response({'error': 'Job bulunamadı'}, status=404)
        return Response({
            'status': job.status,
            'total': job.total,
            'done': job.done,
            'ok': job.ok,
            'failed': job.failed,
            'skipped': job.skipped,
            'log': job.log,
            'errors': job.errors,
        })

    @action(detail=False, methods=['post'], url_path='import-stop')
    def import_stop(self, request):
        job_id = request.data.get('job_id')
        if not job_id:
            return Response({'error': 'job_id gerekli'}, status=400)
        try:
            job = ImportJob.objects.get(id=job_id)
        except ImportJob.DoesNotExist:
            return Response({'error': 'Job bulunamadı'}, status=404)
        job.stop_requested = True
        job.save(update_fields=['stop_requested'])
        return Response({'ok': True})

    @action(detail=False, methods=['post'], url_path='parse-excel')
    def parse_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Dosya gereklidir.'}, status=400)
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Sadece .xlsx dosyaları desteklenir.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            return Response({'error': f'Dosya okunamadı: {str(e)}'}, status=400)
        rows = []
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            rows.append({
                'code': str(row[0]).strip(),
                'address': str(row[1]).strip() if row[1] else '',
            })
        codes = [r['code'] for r in rows]
        existing = {
            e.personnel_code: e.geocode_status
            for e in Employee.objects.filter(personnel_code__in=codes)
        }
        for r in rows:
            r['exists'] = r['code'] in existing
            r['existing_status'] = existing.get(r['code'])
        return Response({'rows': rows})

    @action(detail=False, methods=['post'], url_path='import-row')
    def import_row(self, request):
        personnel_code = str(request.data.get('code', '')).strip()
        address = str(request.data.get('address', '')).strip().lower()
        if not personnel_code:
            logger.warning('import_row: sicil no boş')
            return Response({'status': 'failed', 'reason': 'Sicil no boş'})
        if not address:
            logger.warning('import_row [%s]: adres boş', personnel_code)
            return Response({'status': 'failed', 'reason': 'Adres boş'})
        force = request.data.get('force', False)
        emp, created = Employee.objects.get_or_create(
            personnel_code=personnel_code,
            defaults={'address': address, 'geocode_status': 'pending'},
        )
        if not created and emp.geocode_status == 'ok' and not force:
            logger.debug('import_row [%s]: atlandı (zaten ok)', personnel_code)
            return Response({'status': 'skipped'})
        logger.info('import_row [%s]: geocode isteği gönderiliyor → "%s"', personnel_code, address)
        result = geocode_address(address)
        if result['ok']:
            emp.lat = result['lat']
            emp.lng = result['lng']
            emp.api_address = result['api_address']
            emp.geocode_status = 'ok'
            emp.address = address
            emp.save()
            logger.info('import_row [%s]: OK → lat=%.5f lng=%.5f api_address="%s"',
                        personnel_code, emp.lat, emp.lng, emp.api_address)
            return Response({'status': 'ok', 'api_address': emp.api_address})
        else:
            reason = result.get('reason', 'unknown')
            detail = result.get('detail', '')
            logger.error('import_row [%s]: HATA reason=%s detail=%s', personnel_code, reason, detail)
            emp.geocode_status = 'failed'
            emp.address = address
            emp.save()
            return Response({'status': 'failed', 'reason': reason})

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Dosya gereklidir.'}, status=400)
        if not file.name.endswith('.xlsx'):
            return Response({'error': 'Sadece .xlsx dosyaları desteklenir.'}, status=400)

        try:
            wb = openpyxl.load_workbook(file)
        except Exception as e:
            return Response({'error': f'Dosya okunamadı: {str(e)}'}, status=400)

        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        ok_count = failed_count = skipped_count = 0
        errors = []

        for row in rows:
            if not row or not row[0]:
                continue
            personnel_code = str(row[0]).strip()
            address = str(row[1]).strip() if row[1] else ''

            if not address:
                errors.append({'code': personnel_code, 'reason': 'Adres boş'})
                failed_count += 1
                continue

            emp, created = Employee.objects.get_or_create(
                personnel_code=personnel_code,
                defaults={'address': address, 'geocode_status': 'pending'},
            )

            if not created and emp.geocode_status == 'ok':
                skipped_count += 1
                continue

            result = geocode_address(address)
            if result['ok']:
                emp.lat = result['lat']
                emp.lng = result['lng']
                emp.api_address = result['api_address']
                emp.geocode_status = 'ok'
                emp.address = address
                emp.save()
                ok_count += 1
            else:
                emp.geocode_status = 'failed'
                emp.address = address
                emp.save()
                failed_count += 1
                errors.append({'code': personnel_code, 'reason': result.get('reason', 'unknown')})

            time.sleep(0.15)

        return Response({
            'ok': ok_count,
            'failed': failed_count,
            'skipped': skipped_count,
            'total': ok_count + failed_count + skipped_count,
            'errors': errors[:20],
        })

    @action(detail=False, methods=['post'], url_path='assign')
    def assign(self, request):
        workspace_id = request.data.get('workspace')
        employee_ids = request.data.get('employee_ids', [])
        try:
            workspace = Workspace.objects.get(id=workspace_id)
        except Workspace.DoesNotExist:
            return Response({'error': 'Workspace not found'}, status=404)
        workspace.employees.set(Employee.objects.filter(id__in=employee_ids))
        return Response({'assigned': len(employee_ids)})
